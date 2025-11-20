"""Generate text-based cost model exports for on-device vs cloud runs.

This script writes:
- docs/cost_model_inputs.csv: Inputs sheet (parameters, defaults, notes)
- docs/cost_model_comparison.csv: Comparison sheet (modeled costs and formulas)
- docs/cost_model.csv: Flat summary for CI consumption

Optional: pass ``--xlsx`` to also emit docs/cost_model.xlsx (requires ``openpyxl``).
"""
from __future__ import annotations

import argparse
import csv
import math
import sys
from importlib import util
from pathlib import Path

# Base paths
ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
DOCS.mkdir(exist_ok=True)

# Input parameters
inputs = [
    ("Session volume (per month)", 50_000, "sessions", "Modeled workload size"),
    ("Avg inference duration", 2.0, "minutes per session", "Compute time per session"),
    ("DAL traffic per session", 0.05, "GB", "Payload sent to/from DAL per session"),
    ("Hot-update cadence", 4, "per month", "How often model weights/config ship"),
    ("Update package size", 0.5, "GB", "Average payload per update"),
    ("Device amortization", 120.0, "$ per device per month", "On-device hardware amortization"),
    ("Sessions per device capacity", 10_000, "sessions per month", "Throughput a device can serve"),
    ("Update distribution cost", 0.08, "$ per GB", "Network cost to push updates to devices"),
    ("Cloud inference cost", 0.12, "$ per minute", "GPU/CPU minutes for cloud serving"),
    ("DAL egress cost", 0.09, "$ per GB", "Data-layer traffic pricing"),
    ("Cloud control-plane overhead", 500.0, "$ per month", "Ops + storage for hosted services"),
    ("On-device control-plane overhead", 200.0, "$ per month", "Ops for edge fleet management"),
    ("Cloud update distribution cost", 0.02, "$ per GB", "Distribution for shared cloud replicas"),
]

# Derived metrics (Python side for CSV)
def calculate_costs(params: dict[str, float]) -> dict[str, float]:
    session_volume = params["Session volume (per month)"]
    inference_minutes = params["Avg inference duration"]
    dal_gb = params["DAL traffic per session"]
    hot_updates = params["Hot-update cadence"]
    update_size = params["Update package size"]
    device_amort = params["Device amortization"]
    sessions_per_device = params["Sessions per device capacity"]
    update_distribution_cost = params["Update distribution cost"]
    cloud_inference_cost = params["Cloud inference cost"]
    dal_cost_per_gb = params["DAL egress cost"]
    cloud_overhead = params["Cloud control-plane overhead"]
    device_overhead = params["On-device control-plane overhead"]
    cloud_update_distribution_cost = params["Cloud update distribution cost"]

    device_count = math.ceil(session_volume / sessions_per_device)

    dal_traffic_gb = session_volume * dal_gb
    dal_cost = dal_traffic_gb * dal_cost_per_gb

    # On-device
    on_device_inference_cost = device_count * device_amort
    on_device_update_cost = device_count * hot_updates * update_size * update_distribution_cost
    on_device_total = on_device_inference_cost + dal_cost + on_device_update_cost + device_overhead

    # Cloud
    cloud_inference_total = session_volume * inference_minutes * cloud_inference_cost
    cloud_update_cost = hot_updates * update_size * cloud_update_distribution_cost
    cloud_total = cloud_inference_total + dal_cost + cloud_update_cost + cloud_overhead

    return {
        "device_count": device_count,
        "dal_traffic_gb": dal_traffic_gb,
        "dal_cost": dal_cost,
        "on_device_inference_cost": on_device_inference_cost,
        "on_device_update_cost": on_device_update_cost,
        "on_device_overhead": device_overhead,
        "on_device_total": on_device_total,
        "cloud_inference_cost": cloud_inference_total,
        "cloud_update_cost": cloud_update_cost,
        "cloud_overhead": cloud_overhead,
        "cloud_total": cloud_total,
    }


def build_value_refs(params: list[tuple[str, float, str, str]]) -> dict[str, str]:
    return {name: f"B{idx}" for idx, (name, *_rest) in enumerate(params, start=2)}


def comparison_rows(params: list[tuple[str, float, str, str]]):
    value_cells = build_value_refs(params)

    def v(name: str) -> str:
        return f"Inputs!{value_cells[name]}"

    device_count_formula = f"CEILING({v('Session volume (per month)')}/{v('Sessions per device capacity')},1)"
    dal_traffic_formula = f"{v('Session volume (per month)')}*{v('DAL traffic per session')}"
    dal_cost_formula = f"{dal_traffic_formula}*{v('DAL egress cost')}"

    rows = [
        [
            "Scenario",
            "Inference/Compute",
            "DAL Traffic",
            "Hot Updates",
            "Control-plane Overhead",
            "Total Monthly Cost",
            "Notes",
        ],
        [
            "On-device",
            f"{device_count_formula}*{v('Device amortization')}",
            dal_cost_formula,
            f"{device_count_formula}*{v('Hot-update cadence')}*{v('Update package size')}*{v('Update distribution cost')}",
            v("On-device control-plane overhead"),
            "SUM(B2:E2)",
            "Edge devices run locally; cost scales with fleet size",
        ],
        [
            "Cloud",
            f"{v('Session volume (per month)')}*{v('Avg inference duration')}*{v('Cloud inference cost')}",
            dal_cost_formula,
            f"{v('Hot-update cadence')}*{v('Update package size')}*{v('Cloud update distribution cost')}",
            v("Cloud control-plane overhead"),
            "SUM(B3:E3)",
            "Hosted inference; costs track usage and platform overhead",
        ],
        [],
        ["Derived Metrics", "Value", "Units", "Definition"],
        ["Device count (on-device)", device_count_formula, "devices", "Fleet required for modeled sessions"],
        ["DAL traffic per month", dal_traffic_formula, "GB", "Session volume multiplied by DAL payload"],
    ]

    return rows


def build_workbook(params: list[tuple[str, float, str, str]]):
    spec = util.find_spec("openpyxl")
    if spec is None:
        raise SystemExit("openpyxl is required for --xlsx output; install it or omit the flag")

    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Alignment  # type: ignore

    wb = Workbook()
    wb.remove(wb.active)

    # Inputs sheet
    inputs_ws = wb.create_sheet("Inputs")
    inputs_ws.append(["Parameter", "Value", "Units/Notes", "Description"])
    value_cells = build_value_refs(params)
    for name, value, units, desc in params:
        row = [name, value, units, desc]
        inputs_ws.append(row)
    for col in range(1, 5):
        inputs_ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
    inputs_ws.freeze_panes = "A2"

    # Comparison sheet
    cmp_ws = wb.create_sheet("Comparison")
    cmp_ws.append(["Scenario", "Inference/Compute", "DAL Traffic", "Hot Updates", "Control-plane Overhead", "Total Monthly Cost", "Notes"])

    # Helper formulas
    def v(name: str) -> str:
        return f"Inputs!{value_cells[name]}"

    device_count_formula = f"CEILING({v('Session volume (per month)')}/{v('Sessions per device capacity')},1)"
    dal_traffic_formula = f"{v('Session volume (per month)')}*{v('DAL traffic per session')}"
    dal_cost_formula = f"{dal_traffic_formula}*{v('DAL egress cost')}"

    cmp_ws.append([
        "On-device",
        f"{device_count_formula}*{v('Device amortization')}",
        dal_cost_formula,
        f"{device_count_formula}*{v('Hot-update cadence')}*{v('Update package size')}*{v('Update distribution cost')}",
        v("On-device control-plane overhead"),
        f"SUM(B2:E2)",
        "Edge devices run locally; cost scales with fleet size",
    ])

    cmp_ws.append([
        "Cloud",
        f"{v('Session volume (per month)')}*{v('Avg inference duration')}*{v('Cloud inference cost')}",
        dal_cost_formula,
        f"{v('Hot-update cadence')}*{v('Update package size')}*{v('Cloud update distribution cost')}",
        v("Cloud control-plane overhead"),
        f"SUM(B3:E3)",
        "Hosted inference; costs track usage and platform overhead",
    ])

    # Derived metrics section
    cmp_ws.append([])
    cmp_ws.append(["Derived Metrics", "Value", "Units", "Definition"])
    cmp_ws.append(["Device count (on-device)", device_count_formula, "devices", "Fleet required for modeled sessions"])
    cmp_ws.append(["DAL traffic per month", dal_traffic_formula, "GB", "Session volume multiplied by DAL payload"],)

    for col in range(1, 8):
        cmp_ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
    cmp_ws.freeze_panes = "A2"

    return wb


def export_inputs_csv(params: list[tuple[str, float, str, str]]) -> None:
    inputs_path = DOCS / "cost_model_inputs.csv"
    with inputs_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Parameter", "Value", "Units/Notes", "Description"])
        writer.writerows(params)


def export_comparison_csv(params: list[tuple[str, float, str, str]]) -> None:
    comparison_path = DOCS / "cost_model_comparison.csv"
    with comparison_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        for row in comparison_rows(params):
            writer.writerow(row)


def export_summary_csv(costs: dict[str, float]) -> None:
    csv_path = DOCS / "cost_model.csv"
    with csv_path.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["scenario", "inference_compute", "dal_traffic", "hot_updates", "overhead", "total"])
        writer.writerow([
            "On-device",
            f"{costs['on_device_inference_cost']:.2f}",
            f"{costs['dal_cost']:.2f}",
            f"{costs['on_device_update_cost']:.2f}",
            f"{costs['on_device_overhead']:.2f}",
            f"{costs['on_device_total']:.2f}",
        ])
        writer.writerow([
            "Cloud",
            f"{costs['cloud_inference_cost']:.2f}",
            f"{costs['dal_cost']:.2f}",
            f"{costs['cloud_update_cost']:.2f}",
            f"{costs['cloud_overhead']:.2f}",
            f"{costs['cloud_total']:.2f}",
        ])


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="also write docs/cost_model.xlsx (requires openpyxl)",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    params_dict = {name: value for name, value, _, _ in inputs}
    costs = calculate_costs(params_dict)

    export_inputs_csv(inputs)
    export_comparison_csv(inputs)
    export_summary_csv(costs)

    if args.xlsx:
        wb = build_workbook(inputs)
        xlsx_path = DOCS / "cost_model.xlsx"
        wb.save(xlsx_path)


if __name__ == "__main__":
    main(sys.argv[1:])
